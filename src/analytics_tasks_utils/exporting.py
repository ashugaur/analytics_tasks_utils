# %% Export functions

## Dependencies
from pathlib import Path
import zipfile
import os
from datetime import datetime
import shutil
from analytics_tasks_utils.os_functions import open_file_folder
import pandas as pd
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill


def export_folder_as_zip(source_folder, destination_folder, exclude_folder_names=None):
    if exclude_folder_names is None:
        exclude_folder_names = []

    os.chdir(destination_folder)

    output_filename = str(source_folder).rsplit("\\")[-1] + ".zip"
    zf = zipfile.ZipFile(output_filename, "w", zipfile.ZIP_DEFLATED)

    for dirname, subdirs, files in os.walk(source_folder):
        # Check if current directory or any parent directory should be excluded
        relative_path = os.path.relpath(dirname, source_folder)
        path_parts = relative_path.split(os.sep)

        # Skip if any part of the path matches excluded folder names
        if any(part in exclude_folder_names for part in path_parts):
            subdirs.clear()  # Don't traverse subdirectories of excluded folders
            continue

        # Also check the folder name itself
        folder_name = os.path.basename(dirname)
        if folder_name in exclude_folder_names:
            subdirs.clear()  # Don't traverse subdirectories of excluded folders
            continue

        zf.write(dirname)
        for filename in files:
            zf.write(os.path.join(dirname, filename))

    zf.close()


def export_folder_as_zip_nfp(
    source_folder, destination_folder, exclude_folder_names=None
):
    import os
    import zipfile

    if exclude_folder_names is None:
        exclude_folder_names = []

    os.chdir(destination_folder)

    output_filename = str(source_folder).rsplit("\\")[-1] + ".zip"

    # Create zip file
    with zipfile.ZipFile(output_filename, "w", zipfile.ZIP_DEFLATED) as zf:
        # Walk through the source folder
        for dirname, subdirs, files in os.walk(source_folder):
            # Check if current directory or any parent directory should be excluded
            relative_path = os.path.relpath(dirname, source_folder)
            path_parts = relative_path.split(os.sep)

            # Skip if any part of the path matches excluded folder names
            if any(part in exclude_folder_names for part in path_parts):
                subdirs.clear()  # Don't traverse subdirectories of excluded folders
                continue

            # Also check the folder name itself
            folder_name = os.path.basename(dirname)
            if folder_name in exclude_folder_names:
                subdirs.clear()  # Don't traverse subdirectories of excluded folders
                continue

            # Add files to zip
            for filename in files:
                # Get the absolute path of the file
                absolute_path = os.path.join(dirname, filename)
                # Create arcname (path within the zip file)
                arcname = os.path.relpath(absolute_path, source_folder)

                # Write file to zip
                zf.write(absolute_path, arcname)


def export_folder_as_zip_timestamp(
    source_folder, destination_folder, exclude_folder_names=None
):
    if exclude_folder_names is None:
        exclude_folder_names = []

    now = datetime.now()
    folder_dt = (
        "{:02d}".format(now.year)
        + "{:02d}".format(now.month)
        + "{:02d}".format(now.day)
    )
    file_dt = (
        "{:02d}".format(now.year)
        + "{:02d}".format(now.month)
        + "{:02d}".format(now.day)
        + "_"
        + "{:02d}".format(now.hour)
        + "{:02d}".format(now.minute)
    )

    os.chdir(destination_folder)

    output_filename = str(source_folder).rsplit("\\")[-1] + "_" + file_dt + ".zip"
    zf = zipfile.ZipFile(output_filename, "w", zipfile.ZIP_DEFLATED)

    for dirname, subdirs, files in os.walk(source_folder):
        # Check if current directory or any parent directory should be excluded
        relative_path = os.path.relpath(dirname, source_folder)
        path_parts = relative_path.split(os.sep)

        # Skip if any part of the path matches excluded folder names
        if any(part in exclude_folder_names for part in path_parts):
            subdirs.clear()  # Don't traverse subdirectories of excluded folders
            continue

        # Also check the folder name itself
        folder_name = os.path.basename(dirname)
        if folder_name in exclude_folder_names:
            subdirs.clear()  # Don't traverse subdirectories of excluded folders
            continue

        zf.write(dirname)
        for filename in files:
            zf.write(os.path.join(dirname, filename))

    zf.close()


def backup_folder_force_old(source_folder, destination_folder):
    import shutil
    import os

    try:
        # Delete the destination folder if it exists
        if os.path.exists(destination_folder):
            shutil.rmtree(destination_folder)

        # Copy the source folder to the destination
        shutil.copytree(source_folder, destination_folder)
        # print(f"\nFolder '{source_folder}'\n\tcopied to \n\t'{destination_folder}' successfully.")
    except shutil.Error as e:
        print(f"Error: {e}")
    except Exception as e:
        print(f"An unexpected error occurred: {e}")


def backup_folder_force(source_folder, destination_folder, exclude_folders=None):
    """
    Copies the source folder to the destination, excluding specified folders.

    Args:
        source_folder (str): The path to the source folder.
        destination_folder (str): The path to the destination folder.
        exclude_folders (list, optional): A list of folder names to exclude. Defaults to None.
    """

    try:
        # Delete the destination folder if it exists
        if os.path.exists(destination_folder):
            shutil.rmtree(destination_folder)

        # Copy the source folder to the destination, excluding specified folders
        if exclude_folders is None:
            shutil.copytree(source_folder, destination_folder)
        else:
            shutil.copytree(
                source_folder,
                destination_folder,
                ignore=shutil.ignore_patterns(*exclude_folders),
            )

    except shutil.Error as e:
        print(f"Error: {e}")
    except Exception as e:
        print(f"An unexpected error occurred: {e}")


def zip_files(file_list, destination_file):
    """
    Zips a list of files into a single zip file.

    Args:
        file_list (list): A list of file paths to zip.
        destination_file (str): The path to the destination zip file.
    """

    try:
        # Create a zip file
        with zipfile.ZipFile(destination_file, "w", zipfile.ZIP_DEFLATED) as zip_file:
            # Add files to the zip file
            for file in file_list:
                if os.path.exists(file):
                    relative_path = os.path.relpath(file)
                    zip_file.write(file, relative_path)
                else:
                    print(f"File not found: {file}")

    except Exception as e:
        print(f"An unexpected error occurred: {e}")


if __name__ == "__main__":
    file_list = [
        "/path/to/file1.txt",
        "/path/to/folder/file2.txt",
        "/path/to/file3.pdf",
    ]
    destination_file = "/path/to/destination/zipfile.zip"

    zip_files(file_list, destination_file)


# %% Dataframe to HTML


def generate_html_with_color_and_copy(hex_color):
    html_copy_button = f'<button onclick="copyToClipboard(\'{hex_color}\')" style="width: 60px; height: 20px; background-color: {hex_color}; border: none;"></button>'
    return f"<div>{html_copy_button}</div>"


def generate_html_from_dataframe(df, color_column_name):
    html_content = """
<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="utf-8">
    <title>Colors.py scan</title>
    <link href="../../assets/data_table/bootstrap.min.css" rel="stylesheet">
    <script src="../../assets/data_table/jquery.min.js"></script>
    <link rel="stylesheet" href="../../assets/data_table/jquery.dataTables.min.css">
    <script type="text/javascript" src="../../assets/data_table/jquery.dataTables.min.js"></script>
    <script type="text/javascript" src="../../assets/data_table/bootstrap.min.js"></script>
    <script>
        $(document).ready(function () {
            $('#myTable').dataTable({
                "pageLength": 100 /*load number of rows*/
            });
        });

        function copyToClipboard(text) {
            navigator.clipboard.writeText(text)
                .then(() => {
                    console.log('Copied to clipboard');
                    showNotification('Copied: ' + text);
                })
                .catch(err => {
                    console.error('Error copying to clipboard:', err);
                });
        }

        function showNotification(message) {
            var notification = document.createElement('div');
            notification.textContent = message;
            notification.style.position = 'fixed';
            notification.style.top = '20px';
            notification.style.left = '50%';
            notification.style.transform = 'translateX(-50%)';
            notification.style.background = '#d95f0e';
            notification.style.padding = '10px';
            notification.style.border = '1px solid #ccc';
            notification.style.borderRadius = '5px';
            notification.style.boxShadow = '0 0 10px rgba(0,0,0,0.1)';
            notification.style.zIndex = '9999';
            document.body.appendChild(notification);
            setTimeout(function() {
                document.body.removeChild(notification);
            }, 3000); // Remove notification after 3 seconds
        }
    </script>
    <style>
        body {
            margin: 0;
            padding: 0;
            /*background-color: transparent;*/ /* Set background color to transparent */
        }
        h1 {
            /* text-align: center; */
            margin-top: 20px;
        }
        table {
            width: 70%;
            margin: 20px auto;
            border-collapse: collapse;
        }
        th, td {
            padding: 8px;
            /* text-align: center; */
            border: 1px solid #ddd;
        }
        th {
            vertical-align: middle;
        }
        .color-cell {
            width: 100px;
            /* text-align: center; */
        }
        button {
            padding: 0;
            line-height: 0;
        }
    </style>
</head>
<body style="margin:20px auto">
    <div class="container">
        <h1 style="padding:0; margin-top:0px"></h1>
        <table id="myTable" class="table table-striped">
            <thead>
                <tr>
"""

    # Add table headers dynamically
    for col in df.columns:
        html_content += f"                    <th>{col}</th>\n"
    html_content += "                    <th>Display color</th>\n"
    html_content += "                </tr>\n"
    html_content += "            </thead>\n"
    html_content += "            <tbody>\n"

    # Iterate over DataFrame rows
    for _, row in df.iterrows():
        html_content += "                <tr>\n"
        for col in df.columns:
            html_content += f"                    <td>{row[col]}</td>\n"
        html_content += f'                    <td class="color-cell">{generate_html_with_color_and_copy(row[color_column_name])}</td>\n'
        html_content += "                </tr>\n"

    html_content += """
            </tbody>
        </table>
    </div>
</body>
</html>
"""

    return html_content


def generate_html_with_color_and_copy_dark(hex_color):
    html_copy_button = f'<button onclick="copyToClipboard(\'{hex_color}\')" style="width: 60px; height: 20px; background-color: {hex_color}; border: none;"></button>'
    return f"<div>{html_copy_button}</div>"


def generate_html_from_dataframe_dark(df, color_column_name):
    html_content = """
<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="utf-8">
    <title>Colors.py scan</title>
    <link href="../../assets/data_table/bootstrap.min.css" rel="stylesheet">
    <script src="../../assets/data_table/jquery.min.js"></script>
    <link rel="stylesheet" href="../../assets/data_table/jquery.dataTables.min.css">
    <script type="text/javascript" src="../../assets/data_table/jquery.dataTables.min.js"></script>
    <script type="text/javascript" src="../../assets/data_table/bootstrap.min.js"></script>
    <script>
        $(document).ready(function () {
            $('#myTable').dataTable({
                "pageLength": 100 /*load number of rows*/
            });
        });

        function copyToClipboard(text) {
            navigator.clipboard.writeText(text)
                .then(() => {
                    console.log('Copied to clipboard');
                    showNotification('Copied: ' + text);
                })
                .catch(err => {
                    console.error('Error copying to clipboard:', err);
                });
        }

        function showNotification(message) {
            var notification = document.createElement('div');
            notification.textContent = message;
            notification.style.position = 'fixed';
            notification.style.top = '20px';
            notification.style.left = '50%';
            notification.style.transform = 'translateX(-50%)';
            notification.style.background = '#d95f0e';
            notification.style.padding = '10px';
            notification.style.border = '1px solid #ccc';
            notification.style.borderRadius = '5px';
            notification.style.boxShadow = '0 0 10px rgba(0,0,0,0.1)';
            notification.style.zIndex = '9999';
            document.body.appendChild(notification);
            setTimeout(function() {
                document.body.removeChild(notification);
            }, 3000); // Remove notification after 3 seconds
        }
    </script>
    <style>
        /* Light mode styles */
        body {
            background-color: white;  /* Background outside the table */
            color: black;
        }

        table {
            background-color: white;
            color: black;
        }

        th, td {
            border: 1px solid #ddd;
            background-color: white;
            color: black;
        }

        /* Dark mode styles */
        @media (prefers-color-scheme: dark) {
            body {
                background-color: #1e2129;  /* Dark mode background outside the table */
                color: #e0e0e0;
            }

            /* Optional: you can also target the container specifically if you want */
            .container {
                background-color: #1e2129; /* Dark mode container background */
            }

            /* Table styles in dark mode */
            table, .dataTables_wrapper {
                background-color: #333 !important;
                color: bdbdbd !important;
            }

            th, td {
                border-color: #555;
                background-color: #444 !important; /* Ensure all table cells have dark background */
                color: #bdbdbd !important; /* Ensure all text in table is white */
            }

            /* DataTable plugin-specific styles */
            .dataTables_wrapper .dataTables_paginate .paginate_button {
                background-color: #444 !important; /* Dark background for pagination buttons */
                color: bdbdbd !important; /* White text on pagination buttons */
            }

            .dataTables_wrapper .dataTables_filter input,
            .dataTables_wrapper .dataTables_length select,
            .dataTables_wrapper .dataTables_info {
                background-color: #444 !important; /* Dark background for inputs and dropdowns */
                color: white !important; /* White text for input fields */
            }

            /* Change hover effect in dark mode */
            tr:hover {
                background-color: #636363 !important;
            }

            /* Fix for "Show entries" and "Search" labels */
            .dataTables_wrapper .dataTables_length label,
            .dataTables_wrapper .dataTables_filter label {
                color: #bdbdbd !important; /* Ensure labels like "Show entries" and "Search" are white */
            }
        }
    </style>
</head>
<body style="margin:20px auto">
    <div class="container">
        <h1 style="padding:0; margin-top:0px"></h1>
        <table id="myTable" class="table table-striped">
            <thead>
                <tr>
"""

    # Add table headers dynamically
    for col in df.columns:
        html_content += f"                    <th>{col}</th>\n"
    html_content += "                    <th>Display color</th>\n"
    html_content += "                </tr>\n"
    html_content += "            </thead>\n"
    html_content += "            <tbody>\n"

    # Iterate over DataFrame rows
    for _, row in df.iterrows():
        html_content += "                <tr>\n"
        for col in df.columns:
            html_content += f"                    <td>{row[col]}</td>\n"
        html_content += f'                    <td class="color-cell">{generate_html_with_color_and_copy_dark(row[color_column_name])}</td>\n'
        html_content += "                </tr>\n"

    html_content += """
            </tbody>
        </table>
    </div>
</body>
</html>
"""

    return html_content


# %% Dataframe to Data table


def dataframe_to_data_table(
    df, func="generate_data_table_from_dataframe_text_dark_internet", out_file=None
):
    func_ref = globals()[func]
    html_content = func_ref(df)
    if out_file is None:
        out_file = Path(
            "C:/my_disk/____tmp/generate_data_table_from_dataframe_text_dark_internet.html"
        )
    with open(Path(out_file), "w") as f:
        f.write(html_content)
    open_file_folder(out_file)


def generate_data_table_from_dataframe_internet(df):
    html_content = """
<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="utf-8">
    <title>DataFrame Table</title>
    <link href="https://cdn.jsdelivr.net/npm/bootstrap@3.4.1/dist/css/bootstrap.min.css" rel="stylesheet">
    <script src="https://code.jquery.com/jquery-3.7.1.min.js"></script>
    <link rel="stylesheet" href="https://cdn.datatables.net/1.13.7/css/jquery.dataTables.min.css">
    <script type="text/javascript" src="https://cdn.datatables.net/1.13.7/js/jquery.dataTables.min.js"></script>
    <script type="text/javascript" src="https://cdn.jsdelivr.net/npm/bootstrap@3.4.1/dist/js/bootstrap.min.js"></script>
    <script>
        $(document).ready(function () {
            $('#myTable').dataTable({
                "pageLength": 100 /*load number of rows*/
            });
        });
    </script>
    <style>
        body {
            margin: 0;
            padding: 0;
            /*background-color: transparent;*/ /* Set background color to transparent */
        }
        h1 {
            /* text-align: center; */
            margin-top: 20px;
        }
        table {
            width: 70%;
            margin: 20px auto;
            border-collapse: collapse;
        }
        th, td {
            padding: 8px;
            /* text-align: center; */
            border: 1px solid #ddd;
        }
        th {
            vertical-align: middle;
        }
        .color-cell {
            width: 100px;
            /* text-align: center; */
        }
        button {
            padding: 0;
            line-height: 0;
        }
    </style>
</head>
<body style="margin:20px auto">
    <div class="container">
        <h1 style="padding:0; margin-top:0px"></h1>
        <table id="myTable" class="table table-striped">
            <thead>
                <tr>
"""

    # Add table headers dynamically
    for col in df.columns:
        html_content += f"                    <th>{col}</th>\n"
    html_content += "                </tr>\n"
    html_content += "            </thead>\n"
    html_content += "            <tbody>\n"

    # Iterate over DataFrame rows
    for _, row in df.iterrows():
        html_content += "                <tr>\n"
        for col in df.columns:
            html_content += f"                    <td>{row[col]}</td>\n"
        html_content += "                </tr>\n"

    html_content += """
            </tbody>
        </table>
    </div>
</body>
</html>
"""

    return html_content


def generate_data_table_from_dataframe(df):
    html_content = """
<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="utf-8">
    <title>DataFrame Table</title>
    <link href="../../assets/data_table/bootstrap.min.css" rel="stylesheet">
    <script src="../../assets/data_table/jquery.min.js"></script>
    <link rel="stylesheet" href="../../assets/data_table/jquery.dataTables.min.css">
    <script type="text/javascript" src="../../assets/data_table/jquery.dataTables.min.js"></script>
    <script type="text/javascript" src="../../assets/data_table/bootstrap.min.js"></script>
    <script>
        $(document).ready(function () {
            $('#myTable').dataTable({
                "pageLength": 100 /*load number of rows*/
            });
        });
    </script>
    <style>
        body {
            margin: 0;
            padding: 0;
            /*background-color: transparent;*/ /* Set background color to transparent */
        }
        h1 {
            /* text-align: center; */
            margin-top: 20px;
        }
        table {
            width: 70%;
            margin: 20px auto;
            border-collapse: collapse;
        }
        th, td {
            padding: 8px;
            /* text-align: center; */
            border: 1px solid #ddd;
        }
        th {
            vertical-align: middle;
        }
        .color-cell {
            width: 100px;
            /* text-align: center; */
        }
        button {
            padding: 0;
            line-height: 0;
        }
    </style>
</head>
<body style="margin:20px auto">
    <div class="container">
        <h1 style="padding:0; margin-top:0px"></h1>
        <table id="myTable" class="table table-striped">
            <thead>
                <tr>
"""

    # Add table headers dynamically
    for col in df.columns:
        html_content += f"                    <th>{col}</th>\n"
    html_content += "                </tr>\n"
    html_content += "            </thead>\n"
    html_content += "            <tbody>\n"

    # Iterate over DataFrame rows
    for _, row in df.iterrows():
        html_content += "                <tr>\n"
        for col in df.columns:
            html_content += f"                    <td>{row[col]}</td>\n"
        html_content += "                </tr>\n"

    html_content += """
            </tbody>
        </table>
    </div>
</body>
</html>
"""

    return html_content


def generate_data_table_from_dataframe_dark_internet(df):
    html_content = """
<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="utf-8">
    <title>DataFrame Table</title>
    <link href="https://cdn.jsdelivr.net/npm/bootstrap@3.4.1/dist/css/bootstrap.min.css" rel="stylesheet">
    <script src="https://code.jquery.com/jquery-3.7.1.min.js"></script>
    <link rel="stylesheet" href="https://cdn.datatables.net/1.13.7/css/jquery.dataTables.min.css">
    <script type="text/javascript" src="https://cdn.datatables.net/1.13.7/js/jquery.dataTables.min.js"></script>
    <script type="text/javascript" src="https://cdn.jsdelivr.net/npm/bootstrap@3.4.1/dist/js/bootstrap.min.js"></script>
    <script>
        $(document).ready(function () {
            $('#myTable').dataTable({
                "pageLength": 100 /*load number of rows*/
            });
        });
    </script>
    <style>
        /* Light mode styles */
        body {
            background-color: white;  /* Background outside the table */
            color: black;
        }

        table {
            background-color: white;
            color: black;
        }

        th, td {
            border: 1px solid #ddd;
            background-color: white;
            color: black;
        }

        /* Dark mode styles */
        @media (prefers-color-scheme: dark) {
            body {
                background-color: #1e2129;  /* Dark mode background outside the table */
                color: #e0e0e0;
            }

            /* Optional: you can also target the container specifically if you want */
            .container {
                background-color: #1e2129; /* Dark mode container background */
            }

            /* Table styles in dark mode */
            table, .dataTables_wrapper {
                background-color: #333 !important;
                color: bdbdbd !important;
            }

            th, td {
                border-color: #555;
                background-color: #444 !important; /* Ensure all table cells have dark background */
                color: #bdbdbd !important; /* Ensure all text in table is white */
            }

            /* DataTable plugin-specific styles */
            .dataTables_wrapper .dataTables_paginate .paginate_button {
                background-color: #444 !important; /* Dark background for pagination buttons */
                color: bdbdbd !important; /* White text on pagination buttons */
            }

            .dataTables_wrapper .dataTables_filter input,
            .dataTables_wrapper .dataTables_length select,
            .dataTables_wrapper .dataTables_info {
                background-color: #444 !important; /* Dark background for inputs and dropdowns */
                color: white !important; /* White text for input fields */
            }

            /* Change hover effect in dark mode */
            tr:hover {
                background-color: #636363 !important;
            }

            /* Fix for "Show entries" and "Search" labels */
            .dataTables_wrapper .dataTables_length label,
            .dataTables_wrapper .dataTables_filter label {
                color: #bdbdbd !important; /* Ensure labels like "Show entries" and "Search" are white */
            }
        }
    </style>
</head>
<body style="margin:20px auto">
    <div class="container">
        <h1 style="padding:0; margin-top:0px"></h1>
        <table id="myTable" class="table table-striped">
            <thead>
                <tr>
"""

    # Add table headers dynamically
    for col in df.columns:
        html_content += f"                    <th>{col}</th>\n"
    html_content += "                </tr>\n"
    html_content += "            </thead>\n"
    html_content += "            <tbody>\n"

    # Iterate over DataFrame rows
    for _, row in df.iterrows():
        html_content += "                <tr>\n"
        for col in df.columns:
            html_content += f"                    <td>{row[col]}</td>\n"
        html_content += "                </tr>\n"

    html_content += """
            </tbody>
        </table>
    </div>
</body>
</html>
"""

    return html_content


def generate_data_table_from_dataframe_dark(df):
    html_content = """
<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="utf-8">
    <title>DataFrame Table</title>
    <link href="../../assets/data_table/bootstrap.min.css" rel="stylesheet">
    <script src="../../assets/data_table/jquery.min.js"></script>
    <link rel="stylesheet" href="../../assets/data_table/jquery.dataTables.min.css">
    <script type="text/javascript" src="../../assets/data_table/jquery.dataTables.min.js"></script>
    <script type="text/javascript" src="../../assets/data_table/bootstrap.min.js"></script>
    <script>
        $(document).ready(function () {
            $('#myTable').dataTable({
                "pageLength": 100 /*load number of rows*/
            });
        });
    </script>
    <style>
        /* Light mode styles */
        body {
            background-color: white;  /* Background outside the table */
            color: black;
        }

        table {
            background-color: white;
            color: black;
        }

        th, td {
            border: 1px solid #ddd;
            background-color: white;
            color: black;
        }

        /* Dark mode styles */
        @media (prefers-color-scheme: dark) {
            body {
                background-color: #1e2129;  /* Dark mode background outside the table */
                color: #e0e0e0;
            }

            /* Optional: you can also target the container specifically if you want */
            .container {
                background-color: #1e2129; /* Dark mode container background */
            }

            /* Table styles in dark mode */
            table, .dataTables_wrapper {
                background-color: #333 !important;
                color: bdbdbd !important;
            }

            th, td {
                border-color: #555;
                background-color: #444 !important; /* Ensure all table cells have dark background */
                color: #bdbdbd !important; /* Ensure all text in table is white */
            }

            /* DataTable plugin-specific styles */
            .dataTables_wrapper .dataTables_paginate .paginate_button {
                background-color: #444 !important; /* Dark background for pagination buttons */
                color: bdbdbd !important; /* White text on pagination buttons */
            }

            .dataTables_wrapper .dataTables_filter input,
            .dataTables_wrapper .dataTables_length select,
            .dataTables_wrapper .dataTables_info {
                background-color: #444 !important; /* Dark background for inputs and dropdowns */
                color: white !important; /* White text for input fields */
            }

            /* Change hover effect in dark mode */
            tr:hover {
                background-color: #636363 !important;
            }

            /* Fix for "Show entries" and "Search" labels */
            .dataTables_wrapper .dataTables_length label,
            .dataTables_wrapper .dataTables_filter label {
                color: #bdbdbd !important; /* Ensure labels like "Show entries" and "Search" are white */
            }
        }
    </style>
</head>
<body style="margin:20px auto">
    <div class="container">
        <h1 style="padding:0; margin-top:0px"></h1>
        <table id="myTable" class="table table-striped">
            <thead>
                <tr>
"""

    # Add table headers dynamically
    for col in df.columns:
        html_content += f"                    <th>{col}</th>\n"
    html_content += "                </tr>\n"
    html_content += "            </thead>\n"
    html_content += "            <tbody>\n"

    # Iterate over DataFrame rows
    for _, row in df.iterrows():
        html_content += "                <tr>\n"
        for col in df.columns:
            html_content += f"                    <td>{row[col]}</td>\n"
        html_content += "                </tr>\n"

    html_content += """
            </tbody>
        </table>
    </div>
</body>
</html>
"""

    return html_content


def generate_data_table_from_dataframe_text_dark_internet(df):
    html_content = """
<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="utf-8">
    <title>DataFrame Table</title>
    <link href="https://cdn.jsdelivr.net/npm/bootstrap@3.4.1/dist/css/bootstrap.min.css" rel="stylesheet">
    <script src="https://code.jquery.com/jquery-3.7.1.min.js"></script>
    <link rel="stylesheet" href="https://cdn.datatables.net/1.13.7/css/jquery.dataTables.min.css">
    <script type="text/javascript" src="https://cdn.datatables.net/1.13.7/js/jquery.dataTables.min.js"></script>
    <script type="text/javascript" src="https://cdn.jsdelivr.net/npm/bootstrap@3.4.1/dist/js/bootstrap.min.js"></script>
    <script>
        $(document).ready(function () {
            $('#myTable').dataTable({
                "pageLength": 100 /*load number of rows*/
            });
        });
    </script>
    <style>
        /* Light mode styles */
        body {
            background-color: white;  /* Background outside the table */
            color: black;
        }

        table {
            background-color: white;
            color: black;
        }

        th, td {
            border: 1px solid #ddd;
            background-color: white;
            color: black;
        }

        /* Dark mode styles */
        @media (prefers-color-scheme: dark) {
            body {
                background-color: #1e2129;  /* Dark mode background outside the table */
                color: #e0e0e0;
            }

            /* Optional: you can also target the container specifically if you want */
            .container {
                background-color: #1e2129; /* Dark mode container background */
            }

            /* Table styles in dark mode */
            table, .dataTables_wrapper {
                background-color: #333 !important;
                color: bdbdbd !important;
            }

            th, td {
                border-color: #555;
                background-color: #444 !important; /* Ensure all table cells have dark background */
                color: #bdbdbd !important; /* Ensure all text in table is white */
            }

            /* DataTable plugin-specific styles */
            .dataTables_wrapper .dataTables_paginate .paginate_button {
                background-color: #444 !important; /* Dark background for pagination buttons */
                color: bdbdbd !important; /* White text on pagination buttons */
            }

            .dataTables_wrapper .dataTables_filter input,
            .dataTables_wrapper .dataTables_length select,
            .dataTables_wrapper .dataTables_info {
                background-color: #444 !important; /* Dark background for inputs and dropdowns */
                color: white !important; /* White text for input fields */
            }

            /* Change hover effect in dark mode */
            tr:hover {
                background-color: #636363 !important;
            }

            /* Fix for "Show entries" and "Search" labels */
            .dataTables_wrapper .dataTables_length label,
            .dataTables_wrapper .dataTables_filter label {
                color: #bdbdbd !important; /* Ensure labels like "Show entries" and "Search" are white */
            }
        }
    </style>
</head>
<body style="margin:20px auto">
    <div class="container">
        <h1 style="padding:0; margin-top:0px"></h1>
        <table id="myTable" class="table table-striped">
            <thead>
                <tr>
"""

    # Add table headers dynamically
    for col in df.columns:
        html_content += f"                    <th>{col}</th>\n"
    html_content += "                </tr>\n"
    html_content += "            </thead>\n"
    html_content += "            <tbody>\n"

    # Iterate over DataFrame rows
    for _, row in df.iterrows():
        html_content += "                <tr>\n"
        for col in df.columns:
            html_content += f"                    <td>{row[col]}</td>\n"
        html_content += "                </tr>\n"

    html_content += """
            </tbody>
        </table>
    </div>
</body>
</html>
"""

    return html_content


def generate_data_table_from_dataframe_text_dark(df):
    html_content = """
<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="utf-8">
    <title>DataFrame Table</title>
    <link href="../../../../assets/data_table/bootstrap.min.css" rel="stylesheet">
    <script src="../../../../assets/data_table/jquery.min.js"></script>
    <link rel="stylesheet" href="../../../../assets/data_table/jquery.dataTables.min.css">
    <script type="text/javascript" src="../../../../assets/data_table/jquery.dataTables.min.js"></script>
    <script type="text/javascript" src="../../../../assets/data_table/bootstrap.min.js"></script>
    <script>
        $(document).ready(function () {
            $('#myTable').dataTable({
                "pageLength": 100 /*load number of rows*/
            });
        });
    </script>
    <style>
        /* Light mode styles */
        body {
            background-color: white;  /* Background outside the table */
            color: black;
        }

        table {
            background-color: white;
            color: black;
        }

        th, td {
            border: 1px solid #ddd;
            background-color: white;
            color: black;
        }

        /* Dark mode styles */
        @media (prefers-color-scheme: dark) {
            body {
                background-color: #1e2129;  /* Dark mode background outside the table */
                color: #e0e0e0;
            }

            /* Optional: you can also target the container specifically if you want */
            .container {
                background-color: #1e2129; /* Dark mode container background */
            }

            /* Table styles in dark mode */
            table, .dataTables_wrapper {
                background-color: #333 !important;
                color: bdbdbd !important;
            }

            th, td {
                border-color: #555;
                background-color: #444 !important; /* Ensure all table cells have dark background */
                color: #bdbdbd !important; /* Ensure all text in table is white */
            }

            /* DataTable plugin-specific styles */
            .dataTables_wrapper .dataTables_paginate .paginate_button {
                background-color: #444 !important; /* Dark background for pagination buttons */
                color: bdbdbd !important; /* White text on pagination buttons */
            }

            .dataTables_wrapper .dataTables_filter input,
            .dataTables_wrapper .dataTables_length select,
            .dataTables_wrapper .dataTables_info {
                background-color: #444 !important; /* Dark background for inputs and dropdowns */
                color: white !important; /* White text for input fields */
            }

            /* Change hover effect in dark mode */
            tr:hover {
                background-color: #636363 !important;
            }

            /* Fix for "Show entries" and "Search" labels */
            .dataTables_wrapper .dataTables_length label,
            .dataTables_wrapper .dataTables_filter label {
                color: #bdbdbd !important; /* Ensure labels like "Show entries" and "Search" are white */
            }
        }
    </style>
</head>
<body style="margin:20px auto">
    <div class="container">
        <h1 style="padding:0; margin-top:0px"></h1>
        <table id="myTable" class="table table-striped">
            <thead>
                <tr>
"""

    # Add table headers dynamically
    for col in df.columns:
        html_content += f"                    <th>{col}</th>\n"
    html_content += "                </tr>\n"
    html_content += "            </thead>\n"
    html_content += "            <tbody>\n"

    # Iterate over DataFrame rows
    for _, row in df.iterrows():
        html_content += "                <tr>\n"
        for col in df.columns:
            html_content += f"                    <td>{row[col]}</td>\n"
        html_content += "                </tr>\n"

    html_content += """
            </tbody>
        </table>
    </div>
</body>
</html>
"""

    return html_content


# %% Dataframe to excel


def dataframe_to_excel_no_formatting(
    df,
    out_file=None,
    sheet_name="df",
    index=False,
    open_file=None,
    start_row=1,
    page_bg_color="#E0C9A6",
    header_bg_color="#D4BC96",
    data_bg_color="#D4BC96",
    border_color="#50ABCD",
    header_bold=True,
    alignment="left",
    data_bars=None,
    report_headers=None,
):
    """
    Export DataFrame to Excel with customizable formatting.

    Parameters:
    -----------
    df : DataFrame
        The DataFrame to export
    out_file : str or Path, optional
        Output file path. Defaults to a temp location if None
    sheet_name : str, default "df"
        Name of the Excel sheet
    index : bool, default False
        Whether to include the DataFrame index
    open_file : bool, optional
        Whether to open the file after creation
    start_row : int, default 1
        Starting row for the data (0-indexed). Headers will be at start_row, data at start_row+1
    page_bg_color : str, default "#E0C9A6"
        Background color for the entire page
    header_bg_color : str, default "#D4BC96"
        Background color for header cells
    data_bg_color : str, default "#D4BC96"
        Background color for data cells
    border_color : str, default "#50ABCD"
        Color for cell borders
    header_bold : bool, default True
        Whether to make header text bold
    alignment : str, default "left"
        Text alignment for cells ("left", "center", "right")
    data_bars : dict, optional
        Dictionary mapping column names to data bar colors for conditional formatting.
        Format: {'column_name': 'color'} or {'column_name': {'min_color': '#...', 'max_color': '#...'}}
        Examples:
        - {'Sales': '#63C384'} - Single color gradient (light to dark)
        - {'Revenue': {'min_color': '#FFFFFF', 'max_color': '#FF0000'}} - Two-color gradient
    report_headers : dict, optional
        Dictionary or list of dictionaries for report header rows above the data table.
        Format as dict: {'A1': 'Report name', 'B1': 'First report', 'A2': 'Date', 'B2': '15-11-2025'}
        Format as list: [
            {'cell': 'A1', 'value': 'Report name', 'bold': True, 'font_size': 14},
            {'cell': 'B1', 'value': 'First report'},
            {'cell': 'A2', 'value': 'Date'},
            {'cell': 'B2', 'value': '15-11-2025'}
        ]
        List format supports additional styling: 'bold', 'font_size', 'color', 'bg_color'
    """
    if out_file is None:
        out_file = Path("C:/my_disk/____tmp/dataframe_to_excel_no_formatting.xlsx")

    out_file = Path(out_file)
    file_exists = out_file.exists()

    if file_exists:
        # Load existing workbook and add new sheet
        wb = openpyxl.load_workbook(out_file)

        # Remove sheet if it already exists
        if sheet_name in wb.sheetnames:
            del wb[sheet_name]

        ws = wb.create_sheet(sheet_name)

        # Define styles with custom colors
        header_fill = PatternFill(
            start_color=header_bg_color.lstrip("#"),
            end_color=header_bg_color.lstrip("#"),
            fill_type="solid",
        )
        data_fill = PatternFill(
            start_color=data_bg_color.lstrip("#"),
            end_color=data_bg_color.lstrip("#"),
            fill_type="solid",
        )
        page_fill = PatternFill(
            start_color=page_bg_color.lstrip("#"),
            end_color=page_bg_color.lstrip("#"),
            fill_type="solid",
        )
        border_style = Border(
            left=Side(style="thin", color=border_color.lstrip("#")),
            right=Side(style="thin", color=border_color.lstrip("#")),
            top=Side(style="thin", color=border_color.lstrip("#")),
            bottom=Side(style="thin", color=border_color.lstrip("#")),
        )

        # Write report headers if provided
        if report_headers:
            if isinstance(report_headers, dict):
                # Simple dict format: {'A1': 'value', 'B1': 'value'}
                for cell_ref, value in report_headers.items():
                    cell = ws[cell_ref]
                    cell.value = value
                    cell.fill = page_fill
            elif isinstance(report_headers, list):
                # Advanced list format with styling options
                for header_item in report_headers:
                    cell_ref = header_item.get("cell")
                    value = header_item.get("value")
                    cell = ws[cell_ref]
                    cell.value = value

                    # Apply custom styling if provided
                    if header_item.get("bold", False):
                        cell.font = Font(
                            bold=True,
                            size=header_item.get("font_size", 11),
                            color=header_item.get("color", "000000").lstrip("#"),
                        )
                    else:
                        cell.font = Font(
                            size=header_item.get("font_size", 11),
                            color=header_item.get("color", "000000").lstrip("#"),
                        )

                    # Apply background color if specified
                    if "bg_color" in header_item:
                        cell.fill = PatternFill(
                            start_color=header_item["bg_color"].lstrip("#"),
                            end_color=header_item["bg_color"].lstrip("#"),
                            fill_type="solid",
                        )
                    else:
                        cell.fill = page_fill

        # Apply page background
        for row in range(1, start_row + len(df) + 3):
            for col in range(1, len(df.columns) + 2):
                ws.cell(row, col).fill = page_fill

        # Write headers at start_row
        header_row = start_row + 1  # Convert to 1-indexed
        start_col = 2 if index else 1

        for col_num, value in enumerate(df.columns, start=start_col):
            cell = ws.cell(header_row, col_num, value)
            cell.font = Font(bold=header_bold)
            cell.fill = header_fill
            cell.border = border_style
            cell.alignment = Alignment(horizontal=alignment)

        if index:
            cell = ws.cell(header_row, 1, df.index.name if df.index.name else "")
            cell.font = Font(bold=header_bold)
            cell.fill = header_fill
            cell.border = border_style
            cell.alignment = Alignment(horizontal=alignment)

        # Write data starting at start_row + 1
        for row_num, (idx_val, row) in enumerate(df.iterrows(), start=header_row + 1):
            if index:
                # Convert complex types to string for index
                if isinstance(idx_val, (list, tuple, dict)):
                    idx_val = str(idx_val)
                cell = ws.cell(row_num, 1, idx_val)
                cell.fill = data_fill
                cell.border = border_style
                cell.alignment = Alignment(horizontal=alignment)

            for col_num, value in enumerate(row, start=start_col):
                # Convert complex types to string
                if isinstance(value, (list, tuple, dict)):
                    value = str(value)
                cell = ws.cell(row_num, col_num, value)
                cell.fill = data_fill
                cell.border = border_style
                cell.alignment = Alignment(horizontal=alignment)

        # Apply autofilter
        last_data_row = header_row + len(df)
        last_data_col = len(df.columns) + (1 if index else 0)
        ws.auto_filter.ref = f"A{header_row}:{openpyxl.utils.get_column_letter(last_data_col)}{last_data_row}"

        # Freeze top row at start_row
        ws.freeze_panes = f"A{header_row + 1}"

        # Hide gridlines
        ws.sheet_view.showGridLines = False

        # Auto-adjust column widths
        for idx, column in enumerate(df.columns, start=start_col):
            series = df[column]
            max_len = max(series.astype(str).map(len).max(), len(str(column))) + 2
            ws.column_dimensions[openpyxl.utils.get_column_letter(idx)].width = max_len

        if index:
            ws.column_dimensions["A"].width = (
                max(7, len(str(df.index.name)) + 2) if df.index.name else 7
            )

        # Hide columns beyond the last column with data
        last_col_num = len(df.columns) + (1 if index else 0)
        for col_idx in range(last_col_num + 1, 16385):
            col_letter = openpyxl.utils.get_column_letter(col_idx)
            ws.column_dimensions[col_letter].hidden = True

        # Apply data bars conditional formatting if specified
        if data_bars:
            from openpyxl.formatting.rule import DataBarRule

            for col_name, color_spec in data_bars.items():
                if col_name in df.columns:
                    col_idx = list(df.columns).index(col_name) + (1 if index else 0) + 1
                    col_letter = openpyxl.utils.get_column_letter(col_idx)

                    # Determine colors for data bar
                    if isinstance(color_spec, dict):
                        min_color = color_spec.get("min_color", "#FFFFFF").lstrip("#")
                        max_color = color_spec.get("max_color", "#63C384").lstrip("#")
                    else:
                        # Single color - create gradient from light to specified color
                        max_color = color_spec.lstrip("#")
                        # Create a lighter version for min_color
                        min_color = "FFFFFF"

                    # Create data bar rule
                    data_bar_rule = DataBarRule(
                        start_type="min",
                        end_type="max",
                        color=max_color,
                        showValue=True,
                        minLength=0,
                        maxLength=100,
                    )

                    # Apply to data range (excluding header)
                    data_range = f"{col_letter}{header_row + 1}:{col_letter}{header_row + len(df)}"
                    ws.conditional_formatting.add(data_range, data_bar_rule)

        wb.save(out_file)

    else:
        # Create new workbook with xlsxwriter
        with pd.ExcelWriter(out_file, engine="xlsxwriter") as writer:
            df.to_excel(
                writer,
                sheet_name=sheet_name,
                index=index,
                startrow=start_row + 1,
                header=False,
            )

            workbook = writer.book
            worksheet = writer.sheets[sheet_name]

            # Set formats with custom colors
            page_format = workbook.add_format({"bg_color": page_bg_color})

            border_format = workbook.add_format(
                {"border": 1, "border_color": border_color}
            )

            header_format = workbook.add_format(
                {
                    "bold": header_bold,
                    "border": 1,
                    "border_color": border_color,
                    "bg_color": header_bg_color,
                    "align": alignment,
                }
            )

            data_format = workbook.add_format(
                {
                    "bg_color": data_bg_color,
                    "border": 1,
                    "border_color": border_color,
                    "align": alignment,
                }
            )

            # Write report headers if provided
            if report_headers:
                if isinstance(report_headers, dict):
                    # Simple dict format: {'A1': 'value', 'B1': 'value'}
                    for cell_ref, value in report_headers.items():
                        # Parse cell reference (e.g., 'A1' -> row=0, col=0)
                        col_letter = "".join(filter(str.isalpha, cell_ref))
                        row_num = int("".join(filter(str.isdigit, cell_ref))) - 1
                        col_num = (
                            openpyxl.utils.column_index_from_string(col_letter) - 1
                        )

                        header_text_format = workbook.add_format(
                            {"bg_color": page_bg_color}
                        )
                        worksheet.write(row_num, col_num, value, header_text_format)

                elif isinstance(report_headers, list):
                    # Advanced list format with styling options
                    for header_item in report_headers:
                        cell_ref = header_item.get("cell")
                        value = header_item.get("value")

                        # Parse cell reference
                        col_letter = "".join(filter(str.isalpha, cell_ref))
                        row_num = int("".join(filter(str.isdigit, cell_ref))) - 1
                        col_num = (
                            openpyxl.utils.column_index_from_string(col_letter) - 1
                        )

                        # Create custom format for this header
                        format_props = {
                            "bg_color": header_item.get("bg_color", page_bg_color),
                            "font_size": header_item.get("font_size", 11),
                            "font_color": header_item.get("color", "#000000"),
                        }
                        if header_item.get("bold", False):
                            format_props["bold"] = True

                        custom_format = workbook.add_format(format_props)
                        worksheet.write(row_num, col_num, value, custom_format)

            # Apply page format to all cells
            worksheet.conditional_format(
                0,
                0,
                start_row + len(df) + 2,
                len(df.columns) + (1 if index else 0),
                {"type": "formula", "criteria": "TRUE", "format": page_format},
            )

            # Write header with format at start_row
            for col_num, value in enumerate(df.columns):
                worksheet.write(
                    start_row, col_num + (1 if index else 0), value, header_format
                )
            if index:
                worksheet.write(
                    start_row, 0, df.index.name if df.index.name else "", header_format
                )

            # Format data cells starting at start_row + 1
            for row_num, row in df.iterrows():
                excel_row = start_row + 1 + row_num
                if index:
                    # Convert complex types to string for index
                    idx_value = row.name
                    if isinstance(idx_value, (list, tuple, dict)):
                        idx_value = str(idx_value)
                    worksheet.write(excel_row, 0, idx_value, data_format)
                for col_num, value in enumerate(row):
                    # Convert complex types to string
                    if isinstance(value, (list, tuple, dict)):
                        value = str(value)
                    worksheet.write(
                        excel_row, col_num + (1 if index else 0), value, data_format
                    )

            # Apply border format to all cells
            worksheet.conditional_format(
                0,
                0,
                start_row + len(df) + 2,
                len(df.columns) + (1 if index else 0),
                {"type": "formula", "criteria": "TRUE", "format": border_format},
            )

            # Apply filter
            worksheet.autofilter(
                start_row,
                0,
                start_row + len(df),
                len(df.columns) + (1 if index else 0) - 1,
            )

            # Hide columns beyond the last column with data
            last_col_num = len(df.columns) + (1 if index else 0)
            worksheet.set_column(last_col_num, 16383, None, None, {"hidden": True})

            # Freeze top row at start_row + 1
            worksheet.freeze_panes(start_row + 1, 0)

            # Turn off gridlines
            worksheet.hide_gridlines(2)

            # Auto-adjust column width
            for idx, column in enumerate(df.columns):
                series = df[column]
                max_len = max(series.astype(str).map(len).max(), len(str(column))) + 2
                worksheet.set_column(
                    idx + (1 if index else 0), idx + (1 if index else 0), max_len
                )

            if index:
                worksheet.set_column(
                    0, 0, max(7, len(str(df.index.name)) + 2) if df.index.name else 7
                )

            # Apply data bars conditional formatting if specified
            if data_bars:
                for col_name, color_spec in data_bars.items():
                    if col_name in df.columns:
                        col_idx = list(df.columns).index(col_name) + (1 if index else 0)

                        # Determine colors for data bar
                        if isinstance(color_spec, dict):
                            min_color = color_spec.get("min_color", "#FFFFFF")
                            max_color = color_spec.get("max_color", "#63C384")
                        else:
                            # Single color - create gradient from light to specified color
                            min_color = "#FFFFFF"
                            max_color = color_spec

                        # Apply data bar formatting
                        worksheet.conditional_format(
                            start_row + 1,
                            col_idx,
                            start_row + len(df),
                            col_idx,
                            {
                                "type": "data_bar",
                                "min_type": "min",
                                "max_type": "max",
                                "bar_color": max_color,
                                "bar_only": False,
                                "bar_solid": True,
                                "bar_negative_color": min_color,
                                "bar_negative_color_same": False,
                                "bar_negative_border_color": min_color,
                                "bar_border_color": max_color,
                                "bar_direction": "left",
                            },
                        )

    if open_file:
        open_file_folder(out_file)


if __name__ == "__main__":
    df = pd.DataFrame({"Column A": [1, 2, 3, 4]})

    # Basic usage with default formatting
    dataframe_to_excel_no_formatting(df, "output.xlsx")

    # Custom theme with data bars on specific columns
    dataframe_to_excel_no_formatting(
        df,
        "output.xlsx",
        start_row=2,
        page_bg_color="#F0F0F0",
        header_bg_color="#4472C4",
        data_bg_color="#D9E1F2",
        border_color="#2E5C8A",
        data_bars={
            "Sales": "#63C384",  # Green gradient
            "Revenue": "#5B9BD5",  # Blue gradient
            "Profit": {"min_color": "#FFFFFF", "max_color": "#FF0000"},  # White to red
        },
    )

    # Multiple columns with different color gradients
    dataframe_to_excel_no_formatting(
        df,
        "sales_report.xlsx",
        data_bars={
            "Q1_Sales": "#70AD47",  # Green
            "Q2_Sales": "#4472C4",  # Blue
            "Q3_Sales": "#FFC000",  # Orange
            "Q4_Sales": "#C00000",  # Red
            "Total": {
                "min_color": "#E7E6E6",
                "max_color": "#5B9BD5",
            },  # Custom gradient
        },
    )

    dataframe_to_excel_no_formatting(
        df,
        "output.xlsx",
        start_row=10,
        report_headers=[
            {"cell": "A1", "value": "Report name", "bold": True, "font_size": 14},
            {"cell": "B1", "value": "First report", "bold": True, "color": "#0066CC"},
            {"cell": "A2", "value": "Date"},
            {"cell": "B2", "value": "15-11-2025"},
            {"cell": "A3", "value": "Filter", "bold": True},
            {"cell": "B3", "value": "Paid claims"},
            {"cell": "A4", "value": "Note"},
            {"cell": "B4", "value": "No notes as of now", "bg_color": "#FFFFCC"},
        ],
    )






def dataframe_to_excel(
    df,
    out_file=None,
    sheet_name="df",
    index=False,
    open_file=None,
    start_row=1,
    page_bg_color="#E0C9A6",
    header_bg_color="#D4BC96",
    data_bg_color="#D4BC96",
    border_color="#50ABCD",
    header_bold=True,
    alignment="left",
    data_bars=None,
    report_headers=None,
    column_formats=None,
):
    """
    Export DataFrame to Excel with customizable formatting.

    Parameters:
    -----------
    df : DataFrame
        The DataFrame to export
    out_file : str or Path, optional
        Output file path. Defaults to a temp location if None
    sheet_name : str, default "df"
        Name of the Excel sheet
    index : bool, default False
        Whether to include the DataFrame index
    open_file : bool, optional
        Whether to open the file after creation
    start_row : int, default 1
        Starting row for the data (0-indexed). Headers will be at start_row, data at start_row+1
    page_bg_color : str, default "#E0C9A6"
        Background color for the entire page
    header_bg_color : str, default "#D4BC96"
        Background color for header cells
    data_bg_color : str, default "#D4BC96"
        Background color for data cells
    border_color : str, default "#50ABCD"
        Color for cell borders
    header_bold : bool, default True
        Whether to make header text bold
    alignment : str, default "left"
        Text alignment for cells ("left", "center", "right")
    data_bars : dict, optional
        Dictionary mapping column names to data bar colors for conditional formatting.
        Format: {'column_name': 'color'} or {'column_name': {'min_color': '#...', 'max_color': '#...'}}
        Examples:
        - {'Sales': '#63C384'} - Single color gradient (light to dark)
        - {'Revenue': {'min_color': '#FFFFFF', 'max_color': '#FF0000'}} - Two-color gradient
    report_headers : dict, optional
        Dictionary or list of dictionaries for report header rows above the data table.
        Format as dict: {'A1': 'Report name', 'B1': 'First report', 'A2': 'Date', 'B2': '15-11-2025'}
        Format as list: [
            {'cell': 'A1', 'value': 'Report name', 'bold': True, 'font_size': 14},
            {'cell': 'B1', 'value': 'First report'},
            {'cell': 'A2', 'value': 'Date'},
            {'cell': 'B2', 'value': '15-11-2025'}
        ]
        List format supports additional styling: 'bold', 'font_size', 'color', 'bg_color'
    column_formats : dict, optional
        Dictionary mapping column names to Excel number format strings.
        Common formats:
        - '0.0%' - Percentage with 1 decimal (e.g., 25.3%)
        - '0.00%' - Percentage with 2 decimals (e.g., 25.34%)
        - '#,##0' - Comma-separated integer (e.g., 1,234)
        - '#,##0.00' - Comma-separated with 2 decimals (e.g., 1,234.56)
        - '0.000' - Number with 3 decimals (e.g., 1.234)
        - '$#,##0.00' - Currency format (e.g., $1,234.56)
        - '@' - Text format
        Examples:
        - {'Support': '0.0%', 'sum_A': '#,##0', 'Revenue': '$#,##0.00'}
    """
    if out_file is None:
        out_file = Path("C:/my_disk/____tmp/dataframe_to_excel.xlsx")

    out_file = Path(out_file)
    file_exists = out_file.exists()

    if file_exists:
        # Load existing workbook and add new sheet
        wb = openpyxl.load_workbook(out_file)

        # Remove sheet if it already exists
        if sheet_name in wb.sheetnames:
            del wb[sheet_name]

        ws = wb.create_sheet(sheet_name)

        # Define styles with custom colors
        header_fill = PatternFill(
            start_color=header_bg_color.lstrip("#"),
            end_color=header_bg_color.lstrip("#"),
            fill_type="solid",
        )
        data_fill = PatternFill(
            start_color=data_bg_color.lstrip("#"),
            end_color=data_bg_color.lstrip("#"),
            fill_type="solid",
        )
        page_fill = PatternFill(
            start_color=page_bg_color.lstrip("#"),
            end_color=page_bg_color.lstrip("#"),
            fill_type="solid",
        )
        border_style = Border(
            left=Side(style="thin", color=border_color.lstrip("#")),
            right=Side(style="thin", color=border_color.lstrip("#")),
            top=Side(style="thin", color=border_color.lstrip("#")),
            bottom=Side(style="thin", color=border_color.lstrip("#")),
        )

        # Write report headers if provided
        if report_headers:
            if isinstance(report_headers, dict):
                # Simple dict format: {'A1': 'value', 'B1': 'value'}
                for cell_ref, value in report_headers.items():
                    cell = ws[cell_ref]
                    cell.value = value
                    cell.fill = page_fill
            elif isinstance(report_headers, list):
                # Advanced list format with styling options
                for header_item in report_headers:
                    cell_ref = header_item.get("cell")
                    value = header_item.get("value")
                    cell = ws[cell_ref]
                    cell.value = value

                    # Apply custom styling if provided
                    if header_item.get("bold", False):
                        cell.font = Font(
                            bold=True,
                            size=header_item.get("font_size", 11),
                            color=header_item.get("color", "000000").lstrip("#"),
                        )
                    else:
                        cell.font = Font(
                            size=header_item.get("font_size", 11),
                            color=header_item.get("color", "000000").lstrip("#"),
                        )

                    # Apply background color if specified
                    if "bg_color" in header_item:
                        cell.fill = PatternFill(
                            start_color=header_item["bg_color"].lstrip("#"),
                            end_color=header_item["bg_color"].lstrip("#"),
                            fill_type="solid",
                        )
                    else:
                        cell.fill = page_fill

        # Apply page background
        for row in range(1, start_row + len(df) + 3):
            for col in range(1, len(df.columns) + 2):
                ws.cell(row, col).fill = page_fill

        # Write headers at start_row
        header_row = start_row + 1  # Convert to 1-indexed
        start_col = 2 if index else 1

        for col_num, value in enumerate(df.columns, start=start_col):
            cell = ws.cell(header_row, col_num, value)
            cell.font = Font(bold=header_bold)
            cell.fill = header_fill
            cell.border = border_style
            cell.alignment = Alignment(horizontal=alignment)

        if index:
            cell = ws.cell(header_row, 1, df.index.name if df.index.name else "")
            cell.font = Font(bold=header_bold)
            cell.fill = header_fill
            cell.border = border_style
            cell.alignment = Alignment(horizontal=alignment)

        # Write data starting at start_row + 1
        for row_num, (idx_val, row) in enumerate(df.iterrows(), start=header_row + 1):
            if index:
                # Convert complex types to string for index
                if isinstance(idx_val, (list, tuple, dict)):
                    idx_val = str(idx_val)
                cell = ws.cell(row_num, 1, idx_val)
                cell.fill = data_fill
                cell.border = border_style
                cell.alignment = Alignment(horizontal=alignment)

            for col_num, (col_name, value) in enumerate(zip(df.columns, row), start=start_col):
                # Convert complex types to string
                if isinstance(value, (list, tuple, dict)):
                    value = str(value)
                cell = ws.cell(row_num, col_num, value)
                cell.fill = data_fill
                cell.border = border_style
                cell.alignment = Alignment(horizontal=alignment)
                
                # Apply custom number format if specified
                if column_formats and col_name in column_formats:
                    cell.number_format = column_formats[col_name]

        # Apply autofilter
        last_data_row = header_row + len(df)
        last_data_col = len(df.columns) + (1 if index else 0)
        ws.auto_filter.ref = f"A{header_row}:{openpyxl.utils.get_column_letter(last_data_col)}{last_data_row}"

        # Freeze top row at start_row
        ws.freeze_panes = f"A{header_row + 1}"

        # Hide gridlines
        ws.sheet_view.showGridLines = False

        # Auto-adjust column widths
        for idx, column in enumerate(df.columns, start=start_col):
            series = df[column]
            max_len = max(series.astype(str).map(len).max(), len(str(column))) + 2
            ws.column_dimensions[openpyxl.utils.get_column_letter(idx)].width = max_len

        if index:
            ws.column_dimensions["A"].width = (
                max(7, len(str(df.index.name)) + 2) if df.index.name else 7
            )

        # Hide columns beyond the last column with data
        last_col_num = len(df.columns) + (1 if index else 0)
        for col_idx in range(last_col_num + 1, 16385):
            col_letter = openpyxl.utils.get_column_letter(col_idx)
            ws.column_dimensions[col_letter].hidden = True

        # Apply data bars conditional formatting if specified
        if data_bars:
            from openpyxl.formatting.rule import DataBarRule

            for col_name, color_spec in data_bars.items():
                if col_name in df.columns:
                    col_idx = list(df.columns).index(col_name) + (1 if index else 0) + 1
                    col_letter = openpyxl.utils.get_column_letter(col_idx)

                    # Determine colors for data bar
                    if isinstance(color_spec, dict):
                        min_color = color_spec.get("min_color", "#FFFFFF").lstrip("#")
                        max_color = color_spec.get("max_color", "#63C384").lstrip("#")
                    else:
                        # Single color - create gradient from light to specified color
                        max_color = color_spec.lstrip("#")
                        # Create a lighter version for min_color
                        min_color = "FFFFFF"

                    # Create data bar rule
                    data_bar_rule = DataBarRule(
                        start_type="min",
                        end_type="max",
                        color=max_color,
                        showValue=True,
                        minLength=0,
                        maxLength=100,
                    )

                    # Apply to data range (excluding header)
                    data_range = f"{col_letter}{header_row + 1}:{col_letter}{header_row + len(df)}"
                    ws.conditional_formatting.add(data_range, data_bar_rule)

        wb.save(out_file)

    else:
        # Create new workbook with xlsxwriter
        with pd.ExcelWriter(out_file, engine="xlsxwriter") as writer:
            df.to_excel(
                writer,
                sheet_name=sheet_name,
                index=index,
                startrow=start_row + 1,
                header=False,
            )

            workbook = writer.book
            worksheet = writer.sheets[sheet_name]

            # Set formats with custom colors
            page_format = workbook.add_format({"bg_color": page_bg_color})

            border_format = workbook.add_format(
                {"border": 1, "border_color": border_color}
            )

            header_format = workbook.add_format(
                {
                    "bold": header_bold,
                    "border": 1,
                    "border_color": border_color,
                    "bg_color": header_bg_color,
                    "align": alignment,
                }
            )

            data_format = workbook.add_format(
                {
                    "bg_color": data_bg_color,
                    "border": 1,
                    "border_color": border_color,
                    "align": alignment,
                }
            )

            # Create custom formats for columns with specific formatting
            column_format_objects = {}
            if column_formats:
                for col_name, fmt_string in column_formats.items():
                    column_format_objects[col_name] = workbook.add_format(
                        {
                            "bg_color": data_bg_color,
                            "border": 1,
                            "border_color": border_color,
                            "align": alignment,
                            "num_format": fmt_string,
                        }
                    )

            # Write report headers if provided
            if report_headers:
                if isinstance(report_headers, dict):
                    # Simple dict format: {'A1': 'value', 'B1': 'value'}
                    for cell_ref, value in report_headers.items():
                        # Parse cell reference (e.g., 'A1' -> row=0, col=0)
                        col_letter = "".join(filter(str.isalpha, cell_ref))
                        row_num = int("".join(filter(str.isdigit, cell_ref))) - 1
                        col_num = (
                            openpyxl.utils.column_index_from_string(col_letter) - 1
                        )

                        header_text_format = workbook.add_format(
                            {"bg_color": page_bg_color}
                        )
                        worksheet.write(row_num, col_num, value, header_text_format)

                elif isinstance(report_headers, list):
                    # Advanced list format with styling options
                    for header_item in report_headers:
                        cell_ref = header_item.get("cell")
                        value = header_item.get("value")

                        # Parse cell reference
                        col_letter = "".join(filter(str.isalpha, cell_ref))
                        row_num = int("".join(filter(str.isdigit, cell_ref))) - 1
                        col_num = (
                            openpyxl.utils.column_index_from_string(col_letter) - 1
                        )

                        # Create custom format for this header
                        format_props = {
                            "bg_color": header_item.get("bg_color", page_bg_color),
                            "font_size": header_item.get("font_size", 11),
                            "font_color": header_item.get("color", "#000000"),
                        }
                        if header_item.get("bold", False):
                            format_props["bold"] = True

                        custom_format = workbook.add_format(format_props)
                        worksheet.write(row_num, col_num, value, custom_format)

            # Apply page format to all cells
            worksheet.conditional_format(
                0,
                0,
                start_row + len(df) + 2,
                len(df.columns) + (1 if index else 0),
                {"type": "formula", "criteria": "TRUE", "format": page_format},
            )

            # Write header with format at start_row
            for col_num, value in enumerate(df.columns):
                worksheet.write(
                    start_row, col_num + (1 if index else 0), value, header_format
                )
            if index:
                worksheet.write(
                    start_row, 0, df.index.name if df.index.name else "", header_format
                )

            # Format data cells starting at start_row + 1
            for row_num, row in df.iterrows():
                excel_row = start_row + 1 + row_num
                if index:
                    # Convert complex types to string for index
                    idx_value = row.name
                    if isinstance(idx_value, (list, tuple, dict)):
                        idx_value = str(idx_value)
                    worksheet.write(excel_row, 0, idx_value, data_format)
                for col_num, (col_name, value) in enumerate(zip(df.columns, row)):
                    # Convert complex types to string
                    if isinstance(value, (list, tuple, dict)):
                        value = str(value)
                    
                    # Use custom format if specified for this column
                    if column_formats and col_name in column_formats:
                        cell_format = column_format_objects[col_name]
                    else:
                        cell_format = data_format
                    
                    worksheet.write(
                        excel_row, col_num + (1 if index else 0), value, cell_format
                    )

            # Apply border format to all cells
            worksheet.conditional_format(
                0,
                0,
                start_row + len(df) + 2,
                len(df.columns) + (1 if index else 0),
                {"type": "formula", "criteria": "TRUE", "format": border_format},
            )

            # Apply filter
            worksheet.autofilter(
                start_row,
                0,
                start_row + len(df),
                len(df.columns) + (1 if index else 0) - 1,
            )

            # Hide columns beyond the last column with data
            last_col_num = len(df.columns) + (1 if index else 0)
            worksheet.set_column(last_col_num, 16383, None, None, {"hidden": True})

            # Freeze top row at start_row + 1
            worksheet.freeze_panes(start_row + 1, 0)

            # Turn off gridlines
            worksheet.hide_gridlines(2)

            # Auto-adjust column width
            for idx, column in enumerate(df.columns):
                series = df[column]
                max_len = max(series.astype(str).map(len).max(), len(str(column))) + 2
                worksheet.set_column(
                    idx + (1 if index else 0), idx + (1 if index else 0), max_len
                )

            if index:
                worksheet.set_column(
                    0, 0, max(7, len(str(df.index.name)) + 2) if df.index.name else 7
                )

            # Apply data bars conditional formatting if specified
            if data_bars:
                for col_name, color_spec in data_bars.items():
                    if col_name in df.columns:
                        col_idx = list(df.columns).index(col_name) + (1 if index else 0)

                        # Determine colors for data bar
                        if isinstance(color_spec, dict):
                            min_color = color_spec.get("min_color", "#FFFFFF")
                            max_color = color_spec.get("max_color", "#63C384")
                        else:
                            # Single color - create gradient from light to specified color
                            min_color = "#FFFFFF"
                            max_color = color_spec

                        # Apply data bar formatting
                        worksheet.conditional_format(
                            start_row + 1,
                            col_idx,
                            start_row + len(df),
                            col_idx,
                            {
                                "type": "data_bar",
                                "min_type": "min",
                                "max_type": "max",
                                "bar_color": max_color,
                                "bar_only": False,
                                "bar_solid": True,
                                "bar_negative_color": min_color,
                                "bar_negative_color_same": False,
                                "bar_negative_border_color": min_color,
                                "bar_border_color": max_color,
                                "bar_direction": "left",
                            },
                        )

    if open_file:
        open_file_folder(out_file)


if __name__ == "__main__":
    df = pd.DataFrame({
        "Item_A": ["Cardiologist", "Surgeon", "Nephrologist"],
        "sum_A": [2743, 2771, 2718],
        "Support": [0.256, 0.256, 0.2539],
        "Confidence_%": [93.3, 92.4, 93.4],
        "Revenue": [15234.56, 18956.23, 12456.89]
    })

    # Example with column formatting
    dataframe_to_excel(
        df,
        "output_formatted.xlsx",
        column_formats={
            "Support": "0.0%",           # Percentage with 1 decimal
            "sum_A": "#,##0",            # Comma-separated integer
            "Confidence_%": "0.0",       # Number with 1 decimal
            "Revenue": "$#,##0.00"       # Currency format
        }
    )

    # Combined with other features
    dataframe_to_excel(
        df,
        "full_example.xlsx",
        start_row=5,
        page_bg_color="#F0F0F0",
        header_bg_color="#4472C4",
        data_bg_color="#D9E1F2",
        column_formats={
            "sum_A": "#,##0",
            "Support": "0.0%",
            "Confidence": "0.0%",
        },
        data_bars={
            "sum_A": "#63C384"
        },
        report_headers=[
            {"cell": "A1", "value": "Sales Report", "bold": True, "font_size": 14},
            {"cell": "A2", "value": "Date", "bold": True},
            {"cell": "B2", "value": "15-11-2025"}
        ]
    )





